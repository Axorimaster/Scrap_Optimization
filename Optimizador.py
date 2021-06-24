import pandas as pd
import Simplex
import openpyxl as opy
from openpyxl.styles import Alignment
import math
import string
import numpy as np
from decimal import Decimal, ROUND_DOWN


def round_col(df, nom_col, r):
    l_round = df[nom_col].tolist()
    lista = []
    for val in l_round:
        if type(val) == str:
            lista.append(val)
        else:
            lista.append(round(val,r))

    df.drop(columns=[nom_col], inplace=True)
    df[nom_col] = lista


def calc_mix(n_coladas, df_inventario):

    total_scrap = sum(df_inventario['Carga'].tolist())
    df_inventario['Mix'] = (df_inventario['Carga'] / total_scrap) * 100
    df_inventario['Costo'] = ((df_inventario['Carga'] * df_inventario['Precio']) * n_coladas) / 1000000
    df_inventario['Uso'] = (df_inventario['Carga']) * n_coladas
    df_inventario['Volumen'] = df_inventario['Carga'] / df_inventario['Densidad']
    round_col(df_inventario, 'Volumen', 1)
    round_col(df_inventario, 'Mix', 2)
    round_col(df_inventario, 'Costo', 3)
    round_col(df_inventario, 'Uso', 0)

    return(df_inventario)


def calc_nconstrains(df_inventario):

    residual_cons = 0
    var_const = 2*len(df_inventario)
    prod_constrain = 2
    vol_constrain = 1
    n_constrain = residual_cons+prod_constrain+var_const+vol_constrain

    return n_constrain


def calc_contaminantes(carga_eaf, df_inventario, df_comp):

    l_carga = df_inventario['Carga'].tolist()
    l_col = list(df_comp.columns)

    l_cvalue = []
    for col in l_col:
        suma = 0
        l_cont = df_comp[col].tolist()
        for x in range(len(l_carga)):
            suma = suma + l_carga[x] * l_cont[x]

        l_cvalue.append(round(suma / carga_eaf, 3))

    df_cont = pd.DataFrame(list(zip(l_cvalue)), columns=['Valor'], index=l_col)
    return df_cont


def complete_cesta(l_sorted_scrap, l_scrap, l_cesta, l_peso, l_vol_cesta, l_rho, l_sorted_peso, n_base, vol_cesta, carga_cesta, base_scrap_dict):

    x = l_sorted_scrap.index(n_base)
    usindex = l_scrap.index(n_base)

    scrap_val = 0
    scrap_min = base_scrap_dict[n_base][0]
    scrap_max = base_scrap_dict[n_base][1]

    if scrap_max == 0:
        scrap_max = 99

    else:
        scrap_max = scrap_max-scrap_min-1

    while sum(l_cesta) < carga_cesta and l_peso[usindex] > 0 and (sum(l_vol_cesta) + (1 / l_rho[x])) <= vol_cesta and scrap_val <= scrap_max:

        l_peso[usindex] = l_peso[usindex] - 1
        l_sorted_peso[x] = l_sorted_peso[x] - 1
        l_cesta[x] = l_cesta[x] + 1
        l_vol_cesta[x] = l_vol_cesta[x] + (1 / l_rho[x])

        scrap_val += 1
        print(n_base, l_vol_cesta, l_rho[x])


def calc_cesta(num_cestas, carga_cesta, vol_cesta, n_var, df_inventario, base_scrap_dict, fact_cesta):

    global vol_cesta_mod

    def fill_basket(vol_min, load_min):

        load_min_cesta = load_min
        vol_min_cesta = vol_min

        x = 0
        l_cesta = []
        l_vol_cesta = []

        for scrap in l_sorted_scrap:

            usindex = l_scrap.index(scrap)

            if l_sorted_peso[x] <= 0:

                scrap_load = 0

            elif scrap not in l_base_scrap:

                if l_sorted_peso[x] <= (carga_cesta - sum(l_cesta)-load_min_cesta) and ((l_sorted_peso[x]/l_rho_sorted[x]) <= vol_cesta_mod - sum(l_vol_cesta) - vol_min_cesta):

                    scrap_load = l_sorted_peso[x]

                else:

                    scrap_load_carga = carga_cesta - sum(l_cesta) - load_min_cesta
                    scrap_load_vol = round((vol_cesta_mod - sum(l_vol_cesta) - vol_min_cesta)*(l_rho_sorted[x]),0)

                    if scrap_load_carga > scrap_load_vol:
                        scrap_load = scrap_load_vol

                        if scrap_load < 0:
                            scrap_load = 0
                    else:
                        scrap_load = scrap_load_carga

                        if scrap_load < 0:
                            scrap_load = 0

                    if l_sorted_peso[x] - scrap_load <= 1:
                        if scrap == 'LIVIANA':
                            pass
                        else:
                            scrap_load = scrap_load-1

            else:

                if cesta + 1 == num_cestas:

                    scrap_load = l_sorted_peso[x]

                    vol_min_cesta = vol_min_cesta - (base_scrap_dict[scrap][0] / l_rho_sorted[x])
                    load_min_cesta = load_min_cesta - base_scrap_dict[scrap][0]

                else:

                    if l_sorted_peso[x] < base_scrap_dict[scrap][0]:

                        scrap_load = l_sorted_peso[x]

                    else:

                        scrap_load = base_scrap_dict[scrap][0]
                        vol_min_cesta = vol_min_cesta - (base_scrap_dict[scrap][0] / l_rho_sorted[x])
                        load_min_cesta = load_min_cesta - base_scrap_dict[scrap][0]

            scrap_load = round(scrap_load, 0)
            l_sorted_peso[x] = l_sorted_peso[x] - scrap_load
            l_peso[usindex] = l_peso[usindex] - scrap_load
            l_cesta.append(scrap_load)
            l_vol_cesta.append(scrap_load / l_rho_sorted[x])

            x = x + 1

        return l_cesta, l_vol_cesta

    l_index = []
    for x in range(n_var):
        l_index.append(x)
    df_inventario['Index'] = l_index

    l_nom_cestas = []
    l_peso = df_inventario['Carga'].tolist()

    l_scrap = list(df_inventario.index.values)
    df_inventario.sort_values(by=['Energia'], ascending=False, inplace=True)

    l_rho_sorted = df_inventario['Densidad'].tolist()

    l_pot_sorted = df_inventario['Energia'].tolist()

    l_base_scrap = [*base_scrap_dict]
    print("XDDDDDDDDDDDD",l_base_scrap)

    df_base_scrap = df_inventario.loc[l_base_scrap, :]
    df_base_scrap.sort_values(by=['Densidad'], ascending=False, inplace=True)
    l_sorted_base_scrap = list(df_base_scrap.index.values)

    l_sorted_scrap = list(df_inventario.index.values)

    l_sorted_peso = df_inventario['Carga'].tolist()

    l_base_scrap_indx = []

    vol_min = 0

    for scrap in l_base_scrap:

        base_scrap_indx = l_sorted_scrap.index(scrap)
        l_base_scrap_indx.append(base_scrap_indx)

        if l_sorted_peso[base_scrap_indx] == 0:
            base_scrap_dict[scrap][0] = 0

        base_scrap_rho = df_inventario.loc[scrap, 'Densidad']

        vol_min += base_scrap_dict[scrap][0]/base_scrap_rho

    load_min = 0
    for b_sc in base_scrap_dict:
        load_min += base_scrap_dict[b_sc][0]

    l_kpi_cestas = []
    last_cesta = 0

    for cesta in range(num_cestas):

        compressed = False

        if cesta == 0:
            vol_cesta_mod = vol_cesta*0.97
        else:
            vol_cesta_mod = vol_cesta*fact_cesta

        nom_ces = str("Cesta N°" + str(cesta + 1))

        l_cesta, l_vol_cesta = fill_basket(vol_min, load_min)

        pesada_indx = l_sorted_scrap.index('PESADA')
        liviana_index = l_sorted_scrap.index('LIVIANA')

        if l_cesta[pesada_indx] >= 4:

            l_rho_sorted[liviana_index] = 0.6
            l_sorted_peso = df_inventario['Carga'].tolist()
            vol_min = 0
            l_base_scrap_indx = []
            compressed = True

            for scrap in l_base_scrap:

                base_scrap_indx = l_sorted_scrap.index(scrap)
                l_base_scrap_indx.append(base_scrap_indx)

                if l_sorted_peso[base_scrap_indx] == 0:
                    base_scrap_dict[scrap][0] = 0

                base_scrap_rho = df_inventario.loc[scrap, 'Densidad']
                if scrap == 'LIVIANA':
                    base_scrap_rho = 0.6

                vol_min += base_scrap_dict[scrap][0] / base_scrap_rho

            load_min = 0
            for b_sc in base_scrap_dict:
                load_min += base_scrap_dict[b_sc][0]

            l_cesta, l_vol_cesta = fill_basket(vol_min, load_min)

        print(cesta, l_rho_sorted)

        if compressed:
            complete_cesta(l_sorted_scrap, l_scrap, l_cesta, l_peso, l_vol_cesta, l_rho_sorted, l_sorted_peso,
                           'LIVIANA', vol_cesta_mod, carga_cesta, base_scrap_dict)

        for base_scrap in l_sorted_base_scrap:

            complete_cesta(l_sorted_scrap, l_scrap, l_cesta, l_peso, l_vol_cesta, l_rho_sorted, l_sorted_peso, base_scrap, vol_cesta_mod, carga_cesta, base_scrap_dict)

        df_inventario[nom_ces] = l_cesta
        round_col(df_inventario, nom_ces, 0)

        l_nom_cestas.append(nom_ces)
        l_kpi = kpi_cestas(df_inventario[nom_ces].tolist(), l_rho_sorted,l_pot_sorted)
        l_kpi_cestas.append(l_kpi)

        l_rho_sorted = df_inventario['Densidad'].tolist()

        if cesta+1 == num_cestas:
            last_cesta = nom_ces

    l_col = ['Peso [ton]', 'Volumen [m3]', 'Densidad [ton/m3]', 'Energía [kWh]']
    df_cestas = pd.DataFrame(l_kpi_cestas, columns=l_col, index=l_nom_cestas)

    for x in l_col:
        round_col(df_cestas, x, 4)

    df_inventario.sort_values(by=['Index'], inplace=True)
    df_inventario.drop(columns=['Index'], inplace=True)
    err = False

    if df_cestas.at[last_cesta, 'Volumen [m3]'] > vol_cesta_mod:

        err = True

    return df_cestas, err


def calc_potencia(carga_eaf, df_inventario):

    l_potencia_unit = ((df_inventario['Energia'] * df_inventario['Mix'])/100).tolist()
    rend_ef = sum(((df_inventario['Rendimiento'] * df_inventario['Mix'])/100).tolist())

    p_scrap = sum(l_potencia_unit)
    p_escoria = (30 * p_scrap) / 455
    p_total = p_escoria + p_scrap
    e_quim = 3800
    e_el = (carga_eaf * p_total) - e_quim
    pot = 29500
    p_on = 60 / (pot / e_el)
    p_off = 10.5
    ttt = p_on + p_off
    carga_ef = carga_eaf * rend_ef * 0.98
    p = (carga_ef / ttt) * 60

    return round(p, 1), round(ttt, 1), round(p_on, 1), round(p_total), round(e_el)


def gen_report(n_var, df_inventario, file, lkpi):

    l_indicadores = ["Costo", "Densidad", "Rendimiento", "Energía unit.", "Energía total", "Consumo chatarra"]

    l_ind_si = ['$/ton','ton/m3', "%", "kWh/ton", "kWh", "ton"]

    df_indicadores = pd.DataFrame(list(zip(lkpi, l_ind_si)), columns=['Valores', 'Unidades'], index=l_indicadores)

    l_columns = ['Precio [$/ton]', 'Densidad [ton/m3]', 'Inventario [ton/mes]', 'Rendimiento',
                      'Energía [kWh/ton]', 'Carga [ton]', 'Volumen [m3]', 'Mix', 'Costo [MMUSD]', 'Uso [ton]']
    """"
    for x in range(num_cestas):

        nom_cesta = "Cesta N°"+ str(x+1) +" [ton]"
        l_columns.append(nom_cesta)

    df_inventario.columns = l_columns
    """

    workbook = opy.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Reporte'
    worksheet['A1'] = "REPORTE DE SIMULACIÓN"
    worksheet.column_dimensions['A'].width = 20

    """
    worksheet['A37'] = "CONTRACIÓN DE CONTAMINANTES"
    worksheet['A25'] = "DATOS CESTAS"
    """

    l_alphabet = list(string.ascii_uppercase)

    """
    for x in range(n_var):
        worksheet.column_dimensions[l_alphabet[x+1]].width=12
    worksheet.row_dimensions[15].height = 30
    """

    worksheet.sheet_view.showGridLines = False
    workbook.save(file)

    book = opy.load_workbook(file)
    writer = pd.ExcelWriter(file, engine='openpyxl')
    writer.book = book

    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    """
    df_cestas.to_excel(writer, 'Reporte', startrow=25)
    df_contaminantes.to_excel(writer, 'Reporte', startrow=37)
    
    """

    df_indicadores.to_excel(writer, 'Reporte', startrow=3)

    df_inventario = df_inventario.iloc[:, :14]
    df_inventario.to_excel(writer, 'Reporte', startrow=12)

    writer.save()

    workbook = opy.load_workbook(file)
    worksheet = workbook.active

    """
    for i in range(num_cestas + 5):
        letra = l_alphabet[i + 6]
        num_celda = str(16+len(df_inventario))
        nom_celda = str(letra+num_celda)
        num_fin_suma = str(15+len(df_inventario))
        comando = str("=SUM("+letra+"16:"+letra+num_fin_suma+")")
        worksheet[nom_celda] = comando
    """

    """
    for j in range(len(l_columns)):
        letra = l_alphabet[j+1]
        nom_celda = str(letra+"15")
        worksheet[nom_celda].alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
    """

    for j in range(6):
        letra = l_alphabet[j+1]
        nom_celda = str(letra+"26")
        worksheet[nom_celda].alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')

    workbook.save(file)


def check_carga(l_carga, carga_eaf, df_inventario, n_coladas):

    l_inv = df_inventario['Inventario'].tolist()
    l_carga_round = [round(c,0) for c in l_carga]
    l_dec = []

    if sum(l_carga_round) != carga_eaf:

        print("Normalizar carga.")

        if sum(l_carga_round) < carga_eaf:

            for x in l_carga:
                decimal = math.modf(x)

                if decimal[0] <= 0.5 and abs(decimal[0] > 0.0001):

                    l_dec.append(abs(decimal[0]))

                else:
                    l_dec.append(0)

            l_dec.sort(reverse=True)

            eureka = False
            count = 0

            while eureka == False:

                indx = l_dec.index(l_dec[count])
                l_carga_round[indx] = l_carga_round[indx] + 1

                if l_carga_round[indx]*n_coladas >= l_inv[indx]:

                    l_carga_round[indx] = l_carga_round[indx] - 1
                    count = count+1

                else:
                    eureka = True

        elif sum(l_carga_round) > carga_eaf:
            for x in l_carga:

                decimal = math.modf(x)

                if decimal[0] >= 0.5 and abs(decimal[0] > 0.0001):
                    l_dec.append(abs(decimal[0]))
                else:
                    l_dec.append(1)

            min_dec = min(l_dec)
            indx = l_dec.index(min_dec)
            l_carga_round[indx] = l_carga_round[indx] - 1

    df_inventario['Carga'] = l_carga_round


def kpi_cestas(l_cesta,l_rho_sorted, l_pot_sorted):

    peso_cesta = (sum(l_cesta))
    volumen_cesta = sum([a/b for a,b in zip(l_cesta,l_rho_sorted)])

    if volumen_cesta == 0:
        densidad_cesta = 0
    else:
        densidad_cesta = peso_cesta/volumen_cesta

    energia_cesta = sum([round(a,0)*b for a,b in zip(l_cesta,l_pot_sorted)])

    l_kpi = [peso_cesta,volumen_cesta,densidad_cesta,energia_cesta]

    return l_kpi


def main(df_inventario, df_comp, num_cestas, input_inv, l_lifeline, epsilon, base_scrap_dict, carga_eaf, vol_cesta, carga_cesta, n_coladas, fact_cesta):


    try:
        df_inventario = df_inventario.set_index('Nombre')
        df_comp = df_comp.set_index('Nombre')
    except:
        pass

    df_mix = pd.concat([df_inventario,df_comp], axis=1, join='inner')
    df_mix = df_mix[df_mix.Inventario != '0']

    try:
        df_mix.insert(2, 'Lifeline',l_lifeline)
    except:
        pass

    df_mix = np.split(df_mix, [6], axis=1)
    df_inventario = df_mix[0]

    df_comp = df_mix[1]

    for column in df_inventario.columns.tolist():
        df_inventario[column] = df_inventario[column].astype(float)

    for column in df_comp.columns.tolist():
        df_comp[column] = df_comp[column].astype(float)

    df_reserva = df_inventario.copy()

    ##### VARIABLES #####


    prod_mes = input_inv



    ##Setear display del df
    pd.set_option("display.max_rows", None, "display.max_columns", None, "display.max_colwidth", 200)

    ##Lectura hoja de cálculo mix
    n_cons = calc_nconstrains(df_inventario)
    n_var = len(df_inventario)

    ##Formateo de datos de precios e inventarios para Simplex

    l_precios = df_inventario['Precio'].tolist()
    l_precios.append(0)
    l_precios = list(map(float, l_precios))
    l_inv = df_inventario['Inventario'].tolist()
    l_inv = list(map(float,l_inv))
    l_density = df_inventario['Densidad'].tolist()
    l_density = list(map(float, l_density))
    l_rend = df_inventario['Rendimiento'].tolist()
    l_rend = list(map(float, l_rend))
    l_energia = df_inventario['Energia'].tolist()

    l_energia = list(map(float, l_energia))
    l_inv_ll = list(map(lambda x, y: x * y, l_inv, l_lifeline))

    ##Cálculo de número de coladas por mes


    ##Minimización del costo mediante Simplex

    sol_dir, err = Simplex.minimize(n_var, n_cons, l_precios, l_inv_ll, l_density, num_cestas, vol_cesta, carga_eaf, df_comp,
                                    epsilon, n_coladas)

    if err:
        return True

    else:
        if 'min' in sol_dir:
            sol_dir.pop('min')

    l_carga = list(sol_dir.values())
    df_inventario['Carga'] = l_carga


    ##Normalización de los resultados de Simplex

    check_carga(l_carga, carga_eaf, df_inventario, n_coladas)

    ##Formateo de dataframe inventario
    df_inventario = calc_mix(n_coladas, df_inventario)

    ##Cálculo de idicadores de potencia !!!! Suprimido
    steel, ttt, power_on, potencia_unit, energia_el = calc_potencia(carga_eaf, df_inventario)
    l_costo = df_inventario['Costo'].tolist() * 1000000


    print("------------------RESULTADOS------------------")

    l_kpi = []

    ##Cálculo del número de cestas
    Scrap_vol = df_inventario['Volumen'].sum()
    num_load = Scrap_vol / vol_cesta

    ##Cálculo del costo por tonelada de chatarra
    scrap_cost = round(sum(l_costo) /df_inventario['Uso'].sum(), 2)


    l_mix = df_inventario['Mix'].tolist()
    sum_rho = 0
    sum_rend = 0
    sum_energia_unit = 0
    sum_ener = 0

    for x in range(len(l_mix)):
        sum_rho = sum_rho + ((l_mix[x] / 100) * l_density[x])
        sum_rend = sum_rend + ((l_mix[x]/100)*l_rend[x])
        sum_energia_unit = sum_energia_unit + ((l_mix[x]/100)* l_energia[x])
        sum_ener = sum_ener + ((l_carga[x])*l_energia[x])

    rho_mix = round(sum_rho, 2)

    rend_mix = Decimal(sum_rend).quantize(Decimal('.001'), rounding=ROUND_DOWN)

    energia_unit_mix = round(sum_energia_unit, 2)
    energia_mix = round(sum_ener, 0)
    total_scrap = df_inventario['Uso'].sum()


    l_kpi.append(scrap_cost)
    l_kpi.append(rho_mix)
    l_kpi.append(rend_mix*100)
    l_kpi.append(energia_unit_mix)
    l_kpi.append(energia_mix)
    l_kpi.append(total_scrap)


    print("Num cestas:", num_load)
    print("N° de coladas:", round(n_coladas, 0))
    print("Costo chatarra por tonelada de acero:", round(scrap_cost, 2), "$/ton")
    print("Potencia requerida:", potencia_unit, "kWh/ton")
    print("Energía eléctrica consumida:", energia_el, "kWh")
    print("Power On:", power_on, "min")
    print("Tap to tap:", ttt, "min")
    print("Productividad:", steel, "ton/h")
    print("Densidad ponderada:", rho_mix, "ton/m3")
    print("Rendimiento promedio:", rend_mix*100,"%" )

    df_inventario.drop(columns=['Lifeline'], inplace=True)

    ##Cálculo de contaminantes
    print("-----------------CONTAMINANTES----------------")
    df_contaminantes = calc_contaminantes(carga_eaf, df_inventario, df_comp)
    print(df_contaminantes)

    ##Carga de cestas
    print("--------------------CESTAS--------------------")

    df_cestas, err = calc_cesta(num_cestas, carga_cesta, vol_cesta, n_var, df_inventario, base_scrap_dict, fact_cesta)


    if err == True:
        main_err = True
        return main_err

    return df_inventario, df_cestas, l_kpi, df_contaminantes
